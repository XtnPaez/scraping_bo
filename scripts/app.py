import os
import re
import shutil
import webbrowser
import requests
import pdfplumber
import openpyxl
from datetime import datetime

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

BASE_DIR    = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PDF_HOY     = os.path.join(BASE_DIR, "pdf_hoy")
PALABRAS    = os.path.join(BASE_DIR, "palabras.txt")
ARCHIVO     = os.path.join(BASE_DIR, "archivo")
RESULTADOS  = os.path.join(BASE_DIR, "resultados")
REGISTRO    = os.path.join(BASE_DIR, "registro")
EXCEL_PATH  = os.path.join(REGISTRO, "url_monitor.xlsx")
URL_BOLETIN = "https://www.boletinoficial.gob.ar/seccion/primera"

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def encontrar_pdf():
    archivos = [f for f in os.listdir(PDF_HOY) if f.lower().endswith(".pdf")]
    if not archivos:
        print("ERROR: No hay ningun PDF en la carpeta pdf_hoy/")
        return None
    return os.path.join(PDF_HOY, archivos[0])


def leer_palabras():
    with open(PALABRAS, encoding="utf-8") as f:
        palabras = [l.strip() for l in f if l.strip()]
    if not palabras:
        print("ERROR: palabras.txt esta vacio.")
    return palabras


def extraer_parrafos(pdf_path):
    texto_completo = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if texto:
                texto_completo += " " + texto

    texto_completo = re.sub(r'\s+', ' ', texto_completo).strip()
    parrafos_raw = re.split(r'(?<=\.)\s+', texto_completo)
    parrafos = [p.strip() for p in parrafos_raw if len(p.strip()) >= 10]
    return parrafos


def buscar_palabras(parrafos, palabras):
    resultados = {}
    for palabra in palabras:
        patron = re.compile(re.escape(palabra), re.IGNORECASE)
        matches = [p for p in parrafos if patron.search(p)]
        if matches:
            resultados[palabra] = matches
    return resultados


def archivar_pdf(pdf_path, fecha):
    anio = fecha.strftime("%Y")
    mes  = fecha.strftime("%m")
    dia  = fecha.strftime("%d")
    dest = os.path.join(ARCHIVO, anio, mes, dia)
    os.makedirs(dest, exist_ok=True)
    nombre = "boletin_{}.pdf".format(fecha.strftime("%Y-%m-%d"))
    shutil.copy2(pdf_path, os.path.join(dest, nombre))
    print("PDF archivado en: {}/{}".format(dest, nombre))


def borrar_pdf_hoy(pdf_path):
    try:
        os.remove(pdf_path)
        print("PDF eliminado de pdf_hoy/")
    except Exception as e:
        print("ERROR al eliminar PDF de pdf_hoy/: {}".format(e))


def generar_html(resultados, fecha, palabras_buscadas):
    fecha_str = fecha.strftime("%Y-%m-%d")
    os.makedirs(RESULTADOS, exist_ok=True)
    ruta = os.path.join(RESULTADOS, "{}_reporte.html".format(fecha_str))

    total_parrafos = sum(len(v) for v in resultados.values())

    bloques = ""
    for palabra, parrafos in resultados.items():
        items = "".join("<li>{}</li>".format(p) for p in parrafos)
        bloques += """
        <div class="palabra-bloque">
            <h2>🔍 {} <span class="badge">{} parrafo(s)</span></h2>
            <ul>{}</ul>
        </div>
        """.format(palabra, len(parrafos), items)

    html = """<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>scraping_BO — {fecha}</title>
    <style>
        body {{ font-family: Arial, sans-serif; max-width: 960px; margin: 40px auto; padding: 0 20px; background: #f5f5f5; color: #222; }}
        h1 {{ color: #003366; border-bottom: 2px solid #003366; padding-bottom: 8px; }}
        .resumen {{ background: #e8f0fe; border-left: 4px solid #003366; padding: 12px 16px; margin-bottom: 24px; border-radius: 4px; }}
        .palabra-bloque {{ background: white; border: 1px solid #ddd; border-radius: 6px; padding: 16px 20px; margin-bottom: 20px; }}
        h2 {{ color: #003366; font-size: 1.1em; margin-top: 0; }}
        .badge {{ background: #003366; color: white; font-size: 0.75em; padding: 2px 8px; border-radius: 12px; margin-left: 8px; vertical-align: middle; }}
        ul {{ padding-left: 20px; }}
        li {{ margin-bottom: 10px; line-height: 1.6; background: #fafafa; padding: 8px 12px; border-radius: 4px; border-left: 3px solid #ccc; }}
        .footer {{ margin-top: 40px; font-size: 0.8em; color: #888; text-align: center; }}
    </style>
</head>
<body>
    <h1>scraping_BO — Boletin Oficial Primera Seccion</h1>
    <div class="resumen">
        <strong>Fecha:</strong> {fecha} &nbsp;|&nbsp;
        <strong>Palabras buscadas:</strong> {cant_palabras} &nbsp;|&nbsp;
        <strong>Palabras con match:</strong> {cant_matches} &nbsp;|&nbsp;
        <strong>Parrafos encontrados:</strong> {cant_parrafos}
    </div>
    {bloques}
    <div class="footer">Generado automaticamente por scraping_BO</div>
</body>
</html>""".format(
        fecha=fecha_str,
        cant_palabras=len(palabras_buscadas),
        cant_matches=len(resultados),
        cant_parrafos=total_parrafos,
        bloques=bloques
    )

    with open(ruta, "w", encoding="utf-8") as f:
        f.write(html)

    print("Reporte generado: {}".format(ruta))
    return ruta


def monitorear_url(fecha):
    try:
        r = requests.get(URL_BOLETIN, timeout=10)
        disponible = "Si" if r.status_code == 200 else "No"
        print("URL consultada: HTTP {} -> {}".format(r.status_code, disponible))
    except Exception as e:
        disponible = "No"
        print("URL no disponible: {}".format(e))

    try:
        os.makedirs(REGISTRO, exist_ok=True)

        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            ultimo_id = ws.cell(row=ws.max_row, column=1).value or 0
            nuevo_id = int(ultimo_id) + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Monitoreo URL"
            ws.append(["ID", "Fecha", "URL", "Disponible"])
            nuevo_id = 1

        ws.append([nuevo_id, fecha.strftime("%Y-%m-%d"), URL_BOLETIN, disponible])
        wb.save(EXCEL_PATH)
        print("Registrado en url_monitor.xlsx (fila {})".format(nuevo_id))
    except Exception as e:
        print("ERROR al escribir Excel: {}".format(e))


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    fecha = datetime.today()
    print("\n" + "="*60)
    print("  scraping_BO — {}".format(fecha.strftime("%d/%m/%Y")))
    print("="*60 + "\n")

    # 1. Buscar PDF
    pdf_path = encontrar_pdf()
    if not pdf_path:
        monitorear_url(fecha)
        return

    print("PDF encontrado: {}".format(os.path.basename(pdf_path)))

    # 2. Leer palabras
    palabras = leer_palabras()
    if not palabras:
        monitorear_url(fecha)
        return

    print("Palabras a buscar: {}\n".format(len(palabras)))

    # 3. Extraer parrafos
    print("Procesando PDF...")
    parrafos = extraer_parrafos(pdf_path)
    print("Parrafos extraidos: {}".format(len(parrafos)))

    # 4. Buscar
    resultados = buscar_palabras(parrafos, palabras)

    # 5. Resultados
    if resultados:
        total = sum(len(v) for v in resultados.values())
        print("\n MATCHES ENCONTRADOS: {} palabra(s) — {} parrafo(s)\n".format(len(resultados), total))
        for palabra, parrafos_match in resultados.items():
            print("  [{}] -> {} parrafo(s)".format(palabra, len(parrafos_match)))
        archivar_pdf(pdf_path, fecha)
        ruta_reporte = generar_html(resultados, fecha, palabras)
        webbrowser.open("file:///{}".format(ruta_reporte.replace(os.sep, "/")))
    else:
        print("\n SIN RESULTADOS para las palabras buscadas.")
        print("  No se genera reporte ni se archiva el PDF.")

    # 6. Borrar PDF de pdf_hoy/ (siempre)
    print()
    borrar_pdf_hoy(pdf_path)

    # 7. Monitorear URL (siempre)
    print()
    monitorear_url(fecha)

    print("\n" + "="*60 + "\n")


if __name__ == "__main__":
    main()
